"""docstruct 코드 정의서(xlsx)를 만든다.

역할:
    문서 형식별 처리 파이프라인, 모듈 구조, 데이터 모델, 설정을
    시트별로 정리한 정의서를 생성한다.
호출부:
    python make_definition.py
출력:
    /mnt/user-data/outputs/docstruct_정의서.xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
LLM_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, widths: list[int], header_row: int = 1) -> None:
    """시트 공통 서식을 입힌다.

    입력: ws — 시트, widths — 열 너비 목록, header_row — 머리글 행 번호
    출력: 없음
    """
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(widths)):
        for cell in row:
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    for cell in ws[header_row]:
        if cell.column <= len(widths):
            cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
            cell.fill = HEAD_FILL
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def add_rows(ws, header: list[str], rows: list[list]) -> None:
    """머리글과 데이터를 넣는다."""
    ws.append(header)
    for r in rows:
        ws.append(r)




# ──────────────────────────────────────────────── 구조도 그리기 도구
DIAG_COLORS = {
    "입력":   ("FFF2CC", "7F6000"),
    "판별":   ("E2EFDA", "375623"),
    "pdf":    ("DDEBF7", "1F4E79"),
    "hwp":    ("FCE4D6", "833C0C"),
    "hwpx":   ("E4DFEC", "403151"),
    "공통":   ("D9E2F3", "1F3864"),
    "llm":    ("FBE5D6", "C55A11"),
    "출력":   ("D6DCE4", "222A35"),
    "내부":   ("F2F2F2", "595959"),
}


def diag_sheet(wb, title: str, cols: int = 30, col_w: float = 3.2):
    """구조도용 시트를 만든다 (격자 숨김, 좁은 열).

    입력: wb — 워크북, title — 시트 이름, cols — 열 수, col_w — 열 너비
    출력: 시트
    """
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for i in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = col_w
    return ws


def box(ws, row, col, w, h, text, kind="공통", size=11, bold=True, num=None):
    """구조도 상자 하나를 그린다.

    입력:
        row, col  좌상단 위치 (1부터)
        w, h      가로·세로 셀 수
        text      내용
        kind      색 구분 키 (DIAG_COLORS)
        size      글자 크기
        num       순서 번호. 주면 앞에 크게 붙는다
    출력: 없음
    """
    fill_c, font_c = DIAG_COLORS.get(kind, DIAG_COLORS["공통"])
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row + h - 1, end_column=col + w - 1)
    cell = ws.cell(row=row, column=col)
    cell.value = f"{num}   {text}" if num is not None else text
    cell.font = Font(name=FONT, size=size, bold=bold, color=font_c)
    cell.fill = PatternFill("solid", fgColor=fill_c)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    edge = Side(style="medium", color=font_c)
    for r in range(row, row + h):
        for c in range(col, col + w):
            ws.cell(row=r, column=c).border = Border(
                left=edge if c == col else None,
                right=edge if c == col + w - 1 else None,
                top=edge if r == row else None,
                bottom=edge if r == row + h - 1 else None,
            )
    for r in range(row, row + h):
        ws.row_dimensions[r].height = 20


def arrow(ws, row, col, w=1, text="▼", size=14, color="1F3864"):
    """화살표를 놓는다."""
    if w > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + w - 1)
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=True, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18


def label(ws, row, col, w, text, size=10, color="595959", bold=False, align="center"):
    """설명 문구를 놓는다."""
    if w > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + w - 1)
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def title_row(ws, row, text, sub=""):
    """시트 제목."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=26)
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font = Font(name=FONT, size=15, bold=True, color="1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26
    if sub:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=26)
        c = ws.cell(row=row + 1, column=1)
        c.value = sub
        c.font = Font(name=FONT, size=10, color="595959")
        c.alignment = Alignment(horizontal="left", vertical="center")


wb = Workbook()

# ──────────────────────────────────────────────────── 1. 개요
ws = wb.active
ws.title = "1.개요"
ws["A1"] = "docstruct 코드 정의서"
ws["A1"].font = Font(name=FONT, size=16, bold=True, color="1F3864")
ws["A2"] = "HWP / HWPX / PDF 문서 구조화 라이브러리"
ws["A2"].font = Font(name=FONT, size=11, color="595959")

rows = [
    ["항목", "내용"],
    ["패키지명", "docstruct"],
    ["버전", "0.1.46"],
    ["지원 파이썬", "3.10 ~ 3.13"],
    ["입력 형식", ".pdf  /  .hwp  /  .hwpx"],
    ["출력", "document.json · document.md · tables.md · pipeline.md · layout.md · images/ · pages/"],
    ["진입점 (API)", "DocStruct · DocStructBatch · structure() · build_document()"],
    ["진입점 (CLI)", "docstruct <파일|폴더> [옵션]"],
    ["", ""],
    ["핵심 개념", "설명"],
    ["placeholder", "본문에서 표는 <table N>, 그림은 <image N> 으로 치환해 두고 메타를 따로 관리"],
    ["PageTrace", "페이지마다 어떤 모듈이 무엇을 했는지 순서대로 기록 (문제 추적용)"],
    ["content_type", "LLM 이 판정한 실제 유형: table | text | image"],
    ["quality", "표 파싱 품질: sufficient | wrong | insufficient"],
    ["needs_fill", "content_type=table 이고 quality 가 wrong·insufficient 인 표만 재추출"],
    ["LLM 선택성", "표 판정·재추출·그림설명·목차는 모두 선택. 끄면 파싱 결과 그대로 출력"],
]
for r in rows:
    ws.append(r)
for row_idx in (4, 13):
    for cell in ws[row_idx]:
        if cell.value:
            cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
            cell.fill = HEAD_FILL
style_sheet(ws, [22, 96], header_row=4)
for cell in ws[13]:
    if cell.value:
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL


# ──────────────────────────────────────────────── 2. 전체 구조도
ws = diag_sheet(wb, "2.전체 구조도")
title_row(ws, 1, "전체 처리 흐름", "형식에 따라 갈렸다가 다시 합쳐진다. 회색 상자는 선택 단계다.")

box(ws, 4, 10, 8, 2, "입력 파일\n.pdf  .hwp  .hwpx", "입력", size=12)
arrow(ws, 6, 10, 8)
box(ws, 7, 9, 10, 2, "형식 판별\npipeline.source_format()", "판별", size=11, num="1")
arrow(ws, 9, 9, 10)
label(ws, 10, 2, 24, "─────────────  확장자에 따라 갈림  ─────────────", size=9)

box(ws, 12, 2, 7, 2, ".pdf", "pdf", size=13)
box(ws, 12, 11, 7, 2, ".hwp", "hwp", size=13)
box(ws, 12, 20, 7, 2, ".hwpx", "hwpx", size=13)
for c in (2, 11, 20):
    arrow(ws, 14, c, 7)

box(ws, 15, 2, 7, 3, "Docling\n객체인식 → 유형분류\n→ TableFormer → OCR", "pdf", size=10, num="2")
box(ws, 15, 11, 7, 3, "3경로 분기\nhwpml-xml / pyhwp-html\n/ olefile-text", "hwp", size=10, num="2")
box(ws, 15, 20, 7, 3, "python-hwpx\nzip + XML 파싱", "hwpx", size=10, num="2")
label(ws, 18, 2, 7, "→ 3.PDF 흐름도", size=9, color="1F4E79")
label(ws, 18, 11, 7, "→ 4.HWP 흐름도", size=9, color="833C0C")
label(ws, 18, 20, 7, "→ 5.HWPX 흐름도", size=9, color="403151")

for c in (2, 11, 20):
    arrow(ws, 19, c, 7)
label(ws, 20, 2, 24, "─────────────  여기서부터 공통  ─────────────", size=9)

steps = [
    ("3", "PageContent[]  생성\n본문 + <table N> · <image N> placeholder", "공통", 2),
    ("4", "페이지 렌더  media.page_render        [PDF 만]", "내부", 2),
    ("5", "표 판정  tables.assess        [LLM]\ncontent_type · quality 판정", "llm", 2),
    ("6", "표 재추출  tables.fill        [LLM]\nneeds_fill 인 표만 다시 뽑음", "llm", 2),
    ("7", "블록 정규화  tables.tags", "공통", 2),
    ("8", "분할  split.split_document        [split_chars 지정 시]", "내부", 2),
    ("9", "목차 추출  outline.builder        [LLM · --outline]", "llm", 2),
]
r = 22
for num, text, kind, h in steps:
    box(ws, r, 7, 14, h, text, kind, size=10, num=num)
    arrow(ws, r + h, 7, 14)
    r += h + 1

box(ws, r, 7, 14, 3,
    "출력  report\ndocument.json · document.md · tables.md\npipeline.md · layout.md · images/ · pages/",
    "출력", size=10, num="10")

label(ws, r + 4, 2, 12, "■ 선택 단계 (끄면 파싱 결과 그대로)", size=9, color="595959", align="left")
label(ws, r + 5, 2, 12, "■ LLM 필요 — 연결 안 되면 자동 생략", size=9, color="C55A11", align="left")

# ──────────────────────────────────────────────── 3. PDF 흐름도
ws = diag_sheet(wb, "3.PDF 흐름도")
title_row(ws, 1, "PDF 처리 흐름",
          "3단계는 Docling 내부에서 일어난다 — 우리 코드는 옵션을 주고 결과를 받는다.")

box(ws, 4, 9, 10, 2, "보고서.pdf", "입력", size=12)
arrow(ws, 6, 9, 10)
box(ws, 7, 7, 14, 2, "extractors.registry._extract_pdf()\n확장자 → PDF 추출기 매핑", "pdf", size=10, num="1")
arrow(ws, 9, 7, 14)
box(ws, 10, 7, 14, 2, "converters.pdf.docling_backend\n장치 판정 → 옵션 구성 → 컨버터 생성 (1회 캐시)", "pdf", size=10, num="2")
arrow(ws, 12, 7, 14)

box(ws, 13, 3, 22, 1, "▼   Docling 내부 (3단계)   ▼", "내부", size=11)
sub = [
    ("3-1", "객체 인식", "LayoutPredictor (RT-DETR)\n페이지 이미지에서 영역 검출"),
    ("3-2", "유형 분류", "DocItemLabel\nTEXT · TABLE · PICTURE · SECTION_HEADER"),
    ("3-3", "표 구조 인식", "TableFormer\n행·열·병합셀 복원"),
    ("3-4", "문자 인식", "RapidOCR / Tesseract / EasyOCR\n텍스트 레이어 없는 영역만"),
    ("3-5", "그림 설명", "PictureDescriptionApi   [VLM]\n외부 엔드포인트 호출"),
]
r = 15
for num, name, detail in sub:
    box(ws, r, 4, 6, 2, name, "내부", size=11, num=num)
    label(ws, r, 11, 14, detail, size=9, align="left")
    ws.merge_cells(start_row=r, start_column=11, end_row=r + 1, end_column=24)
    if num != "3-5":
        arrow(ws, r + 2, 4, 6, text="▼", size=11, color="595959")
    r += 3

arrow(ws, r, 7, 14)
box(ws, r + 1, 7, 14, 2, "converters.pdf.converter._get_document()\n실패 페이지 수집 · 텍스트 출처 측정", "pdf", size=10, num="4")
arrow(ws, r + 3, 7, 14)
box(ws, r + 4, 7, 14, 2, "extractors.pdf.extract_pdf_pages()\niterate_items() 순회 → 페이지별 본문 조립", "pdf", size=10, num="5")

r2 = r + 7
for num, name, detail in [
    ("5-1", "표 변환", "tables.docling.docling_table_to_markdown()\nTableItem → GFM markdown (다단 헤더 병합)"),
    ("5-2", "그림 저장", "media.picture.save_picture()\nPictureItem → PNG + ImageInfo"),
    ("5-3", "레이아웃 기록", "layout.collect_layout()\n인식 라벨 vs 처리 결과 대조 (진단)"),
]:
    box(ws, r2, 6, 6, 2, name, "내부", size=10, num=num)
    ws.merge_cells(start_row=r2, start_column=13, end_row=r2 + 1, end_column=25)
    label(ws, r2, 13, 13, detail, size=9, align="left")
    r2 += 3

arrow(ws, r2, 7, 14)
box(ws, r2 + 1, 7, 14, 2, "PageContent[]   →   공통 후처리 (6~10)", "공통", size=11)

# ──────────────────────────────────────────────── 4. HWP 흐름도
ws = diag_sheet(wb, "4.HWP 흐름도")
title_row(ws, 1, "HWP 처리 흐름",
          "파일 실체에 따라 세 경로로 갈린다. 아래로 갈수록 얻는 정보가 적다.")

box(ws, 4, 9, 10, 2, "문서.hwp", "입력", size=12)
arrow(ws, 6, 9, 10)
box(ws, 7, 7, 14, 2, "converters.hwp.converter._route()\n파일 실체 판별", "판별", size=10, num="1")
arrow(ws, 9, 7, 14)

paths = [
    ("A", "hwpml-xml", "내용이 실제로는 XML", [
        "hwpml.is_hwpml()  →  XML 확인",
        "hwpml.to_html()  ElementTree 직접 파싱",
        "html.blocks.html_to_markdown()",
    ], "표 구조 보존 · 빠름", "hwp"),
    ("B", "pyhwp-html", "OLE2 바이너리 · hwp5html 성공", [
        "pyhwp.hwp_to_html_str()  CLI 실행 (기본 300초)",
        "pyhwp.assess_pyhwp_html()  결과 검사",
        "html.blocks.html_to_markdown()",
        "converter.extract_bindata_images()  그림 추출",
    ], "표 구조 보존 · 큰 문서는 매우 느림", "hwp"),
    ("C", "olefile-text", "hwp5html 실패·타임아웃 → 최후 수단", [
        "olefile.extract_raw_text()  스트림 직접 해독",
        "└ 인라인 제어문자 건너뛰기 (12바이트 부속)",
        "└ UTF-16 surrogate 쌍 결합",
        "└ 사설영역(PUA) 문자 정리",
        "olefile.clean_text()  문단 구성",
    ], "표 구조 없음 · 재추출 불가", "hwp"),
]
r = 10
for key, name, cond, steps_, note, kind in paths:
    box(ws, r, 2, 4, 2, key, kind, size=16)
    box(ws, r, 7, 9, 2, name, kind, size=12)
    label(ws, r, 17, 9, cond, size=9, align="left")
    ws.merge_cells(start_row=r, start_column=17, end_row=r + 1, end_column=25)
    r += 2
    for st in steps_:
        indent = 8 if st.startswith("└") else 7
        label(ws, r, indent, 18, st, size=9, align="left",
              color="595959" if st.startswith("└") else "222A35")
        ws.row_dimensions[r].height = 16
        r += 1
    label(ws, r, 7, 18, f"→ {note}", size=9, bold=True,
          color="C55A11" if key == "C" else "375623", align="left")
    r += 2

arrow(ws, r, 7, 14)
box(ws, r + 1, 7, 14, 2, "extractors.hwp.extract_hwp_pages()\nPageContent 1개 (페이지 개념 없음)", "공통", size=10, num="2")
arrow(ws, r + 3, 7, 14)
box(ws, r + 4, 7, 14, 2, "split.split_document()   [split_chars 지정 시]\n제N장 · □ · 번호 경계에서 목표 크기까지 모아 분할", "내부", size=10, num="3")
arrow(ws, r + 6, 7, 14)
box(ws, r + 7, 7, 14, 2, "공통 후처리 (표 판정 → 재추출 → 출력)", "공통", size=11)
label(ws, r + 10, 2, 24, "※ 재추출 근거는 원본 <table> HTML. C 경로는 근거가 없어 재추출이 일어나지 않는다.",
      size=9, color="C55A11", align="left")

# ──────────────────────────────────────────────── 5. HWPX 흐름도
ws = diag_sheet(wb, "5.HWPX 흐름도")
title_row(ws, 1, "HWPX 처리 흐름", "zip + XML 구조라 세 형식 중 가장 단순하고 빠르다.")

flow = [
    ("", "문서.hwpx", "입력", 2, ""),
    ("1", "extractors.registry._extract_hwpx()", "hwpx", 2, "확장자 → HWPX 추출기"),
    ("2", "extractors.hwpx.extract_hwpx_pages()", "hwpx", 2, "import 성공 여부를 탐색보다 우선"),
    ("3", "hwpx.HwpxDocument.open()", "내부", 2, "zip 컨테이너 → section*.xml"),
    ("4", "export_rich_markdown()", "내부", 2, "본문·표 → markdown (cellSpan 필요)"),
    ("5", "tables.markdown.inject_table_placeholders()", "hwpx", 2, "markdown 표 → <table N> + TableInfo"),
    ("6", "split.split_document()   [선택]", "내부", 2, "구조 경계 기준 분할"),
    ("7", "공통 후처리 (표 판정 → 재추출 → 출력)", "공통", 2, "재추출 근거는 원본 <table> HTML"),
]
r = 4
for num, text, kind, h, note in flow:
    box(ws, r, 6, 13, h, text, kind, size=11, num=num or None)
    if note:
        ws.merge_cells(start_row=r, start_column=20, end_row=r + 1, end_column=29)
        label(ws, r, 20, 10, note, size=9, align="left")
    if text != flow[-1][1]:
        arrow(ws, r + h, 6, 13)
    r += h + 1

label(ws, r + 1, 2, 26,
      "※ 페이지 이미지가 없어 표 판정·재추출은 원본 HTML 만 근거로 삼는다.",
      size=9, color="595959", align="left")

# ──────────────────────────────────────────────── 6. 계층 구조도
ws = diag_sheet(wb, "6.계층 구조도")
title_row(ws, 1, "모듈 계층", "위 계층이 아래를 참조한다. 역방향 참조와 순환은 없다.")

layers = [
    ("L5", "진입점", ["docstruct.api", "docstruct.cli"], "출력"),
    ("L4", "조립·출력", ["pipeline", "report", "preview", "checks"], "공통"),
    ("L3", "구조화", ["extractors", "tables", "media", "outline", "layout", "split", "models"], "판별"),
    ("L2", "포맷 파싱", ["converters.pdf", "converters.hwp", "converters.html"], "pdf"),
    ("L1", "외부 연동", ["infrastructure.llm.client", "infrastructure.llm.local_vlm"], "llm"),
    ("L0", "설정·플랫폼", ["core.config", "core.winfix"], "입력"),
]
r = 4
for code, name, mods_, kind in layers:
    box(ws, r, 2, 3, 2, code, kind, size=14)
    box(ws, r, 6, 6, 2, name, kind, size=11)
    col = 13
    for m in mods_:
        w = max(4, min(9, len(m) // 2 + 2))
        box(ws, r, col, w, 2, m, "내부", size=9, bold=False)
        col += w + 1
        if col > 40:
            break
    r += 2
    if code != "L0":
        arrow(ws, r, 2, 10, text="▼   아래 계층을 참조", size=10, color="595959")
        r += 1

label(ws, r + 1, 2, 26, "※ tools/verify_package.py 가 계층 위반과 순환 참조를 검사한다.",
      size=9, color="595959", align="left")


# ──────────────────────────────────────────────────── 형식별 경로 요약
ws = wb.create_sheet("7.형식별 경로")
add_rows(ws, [
    "입력 형식", "판별 조건", "경로 이름", "파싱 엔진", "표 구조",
    "그림 추출", "페이지 개념", "재추출 근거",
], [
    [".pdf", "확장자 .pdf", "docling", "Docling (레이아웃모델 + TableFormer + OCR)",
     "보존 (TableFormer)", "O (PictureItem)", "페이지 단위", "페이지 PNG 이미지"],
    [".hwp", "내용이 XML (HWPML)", "hwpml-xml", "ElementTree 직접 파싱",
     "보존 (TABLE 요소)", "X", "없음 — 문서 전체 1장", "원본 <table> HTML"],
    [".hwp", "OLE2 바이너리 + hwp5html 성공", "pyhwp-html", "pyhwp (hwp5html) → BeautifulSoup",
     "보존 (rowspan/colspan)", "O (BinData)", "없음 — 문서 전체 1장", "원본 <table> HTML"],
    [".hwp", "hwp5html 실패·타임아웃", "olefile-text", "olefile 스트림 직접 해독",
     "없음 (텍스트만)", "X", "없음 — 문서 전체 1장", "없음 → 재추출 불가"],
    [".hwpx", "확장자 .hwpx (zip+XML)", "python-hwpx", "python-hwpx export_rich_markdown",
     "보존", "X", "없음 — 문서 전체 1장", "원본 <table> HTML"],
])
style_sheet(ws, [11, 26, 14, 34, 18, 15, 20, 20])

# ──────────────────────────────────────────────────── 3. PDF 파이프라인
ws = wb.create_sheet("8.PDF 단계표")
add_rows(ws, [
    "단계", "구분", "모듈", "함수 / 클래스", "하는 일", "입력", "출력", "LLM", "선택 여부",
], [
    ["1", "진입", "docstruct.pipeline", "build_document()",
     "형식 판별 후 추출기 호출, 단계별 시간 측정", "파일 경로", "PageDocument", "-", "필수"],
    ["2", "라우팅", "docstruct.extractors.registry", "_extract_pdf()",
     "확장자 → PDF 추출기 매핑 (Strategy 패턴)", "경로, image_dir", "ExtractionResult", "-", "필수"],
    ["3", "변환 준비", "converters.pdf.docling_backend", "get_document_converter()",
     "장치 판정(GPU/CPU) → docling 옵션 구성 → DocumentConverter 생성 (1회 캐시)",
     "설정", "DocumentConverter", "-", "필수"],
    ["3-1", "└ 객체 인식", "docling 내부", "LayoutPredictor (RT-DETR)",
     "페이지 이미지에서 영역 검출 — 본문·제목·표·그림·머리말 등",
     "페이지 렌더 이미지", "레이아웃 영역 + 라벨", "-", "필수"],
    ["3-2", "└ 유형 분류", "docling 내부", "DocItemLabel",
     "검출 영역에 라벨 부여 (TEXT / TABLE / PICTURE / SECTION_HEADER …)",
     "레이아웃 영역", "라벨링된 요소", "-", "필수"],
    ["3-3", "└ 표 구조 인식", "docling 내부", "TableFormer",
     "TABLE 영역의 행·열·병합셀 구조 복원", "표 영역 이미지", "TableItem (셀 그리드)", "-",
     "do_table_structure=True"],
    ["3-4", "└ 문자 인식", "docling 내부", "RapidOCR / Tesseract / EasyOCR",
     "텍스트 레이어가 없는 영역(스캔)에 OCR 수행", "영역 이미지", "텍스트", "-",
     "do_ocr=True"],
    ["3-5", "└ 그림 설명", "docling 내부", "PictureDescriptionApi",
     "PICTURE 영역을 외부 VLM 에 보내 캡션 생성", "그림 이미지", "설명 문장", "VLM",
     "선택 (엔드포인트 설정 시)"],
    ["4", "결과 수집", "converters.pdf.converter", "PdfConverter._get_document()",
     "convert() 실행, 페이지 단위 실패 수집, 텍스트 출처(text_layer/ocr) 측정",
     "PDF 경로", "DoclingDocument + failed_pages + page_stats", "-", "필수"],
    ["5", "구조화", "docstruct.extractors.pdf", "extract_pdf_pages()",
     "iterate_items() 로 요소 순회 → 페이지별 본문 조립, 표·그림을 placeholder 로 치환",
     "DoclingDocument", "list[PageContent]", "-", "필수"],
    ["5-1", "└ 표 변환", "docstruct.tables.docling", "docling_table_to_markdown()",
     "TableItem 셀 그리드 → GFM markdown (다단 헤더 병합, span 전파)",
     "TableItem", "markdown 표", "-", "필수"],
    ["5-2", "└ 그림 저장", "docstruct.media.picture", "save_picture()",
     "PictureItem 이미지를 파일로 저장하고 ImageInfo 생성", "PictureItem", "ImageInfo + PNG", "-", "필수"],
    ["5-3", "└ 레이아웃 기록", "docstruct.layout", "collect_layout()",
     "인식 라벨 vs 파이프라인 처리 결과를 대조 기록 (진단용)",
     "DoclingDocument", "list[LayoutItem]", "-", "필수"],
    ["6", "페이지 렌더", "docstruct.media.page_render", "render_page_images()",
     "표가 있는 페이지를 PNG 로 렌더 (표 판정·재추출의 시각 근거)",
     "PDF 경로, 페이지 번호", "PNG 파일", "-", "render_pages=True"],
    ["7", "표 판정", "docstruct.tables.assess", "assess_document()",
     "페이지 이미지 + markdown 을 LLM 에 보내 content_type·quality 판정",
     "PageContent[]", "TableInfo.content_type / quality", "O",
     "assess_tables=True"],
    ["8", "표 재추출", "docstruct.tables.fill", "process_tables()",
     "needs_fill 인 표만 페이지 이미지 근거로 다시 markdown 생성",
     "PageContent[]", "TableInfo.markdown 교체", "O", "fill_tables=True"],
    ["9", "블록 정규화", "docstruct.tables.tags", "normalize_blocks()",
     "<table N> 태그 정리, 이미지로 흡수된 표 제거", "PageContent[]", "PageContent[]", "-", "필수"],
    ["10", "목차 추출", "docstruct.outline.builder", "build_outline()",
     "페이지별 의미 경로(장·절) 추론", "PageDocument", "outline.md", "O", "--outline"],
    ["11", "출력", "docstruct.report", "write_json / write_markdown / …",
     "json + md 5종 저장, placeholder 를 실제 내용으로 펼침", "PageDocument", "파일", "-", "필수"],
])
style_sheet(ws, [6, 13, 27, 27, 44, 20, 26, 6, 18])
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[1].value and row[1].value.startswith("└"):
        for c in row:
            c.fill = SUB_FILL
    if row[7].value in ("O", "VLM"):
        row[7].fill = LLM_FILL

# ──────────────────────────────────────────────────── 4. HWP 파이프라인
ws = wb.create_sheet("9.HWP 단계표")
ws["A1"] = "HWP 는 파일 실체에 따라 세 경로로 갈린다. 판별은 converters.hwp.converter.HwpConverter._route() 가 수행한다."
ws["A1"].font = Font(name=FONT, size=10, bold=True)
ws["A1"].fill = NOTE_FILL
ws.append([])
add_rows(ws, [
    "경로", "단계", "모듈", "함수", "하는 일", "표 구조", "비고",
], [
    ["공통", "1", "docstruct.extractors.registry", "_extract_hwp()",
     "HwpConverter 생성 후 경로 판별", "-", "확장자 .hwp"],
    ["공통", "2", "converters.hwp.converter", "_route()",
     "is_hwpml() → XML 여부 확인 / OLE2 이면 hwp5html 시도", "-",
     "결과: hwpml-xml | pyhwp-html | olefile-text"],

    ["A. hwpml-xml", "3", "converters.hwp.hwpml", "is_hwpml() → to_html()",
     "내용이 실제로는 HWPML(XML) — ElementTree 로 직접 파싱", "보존",
     "TABLE 서브트리 노드를 제외해 본문 중복 방지"],
    ["A. hwpml-xml", "4", "converters.html.blocks", "html_to_markdown()",
     "BeautifulSoup 으로 표 구조 유지하며 markdown 변환", "보존", "다단 헤더 병합 포함"],

    ["B. pyhwp-html", "3", "converters.hwp.pyhwp", "hwp_to_html_str()",
     "hwp5html CLI 실행 (기본 300초 제한, DOCSTRUCT_HWP_TIMEOUT 로 조정)", "보존",
     "큰 문서는 매우 느림 — 3.4MB 가 수 분 초과 가능"],
    ["B. pyhwp-html", "4", "converters.hwp.pyhwp", "assess_pyhwp_html()",
     "변환 결과가 쓸 만한지 검사 (내용 없음·깨짐 판정)", "-", "실패 시 C 경로로"],
    ["B. pyhwp-html", "5", "converters.html.blocks", "html_to_markdown()",
     "표의 rowspan/colspan 을 살려 markdown 변환", "보존", "-"],
    ["B. pyhwp-html", "6", "converters.hwp.converter", "extract_bindata_images()",
     "OLE BinData 스트림에서 그림 추출", "-", "PNG/JPG 저장"],

    ["C. olefile-text", "3", "converters.hwp.olefile", "extract_raw_text()",
     "OLE 스트림을 직접 해독 — 레코드 순회, zlib 해제, UTF-16 디코딩", "없음",
     "hwp5html 실패·타임아웃 시 최후 수단"],
    ["C. olefile-text", "3-1", "converters.hwp.olefile", "└ 인라인 제어문자 처리",
     "코드 1~23 뒤 12바이트 부속 데이터 건너뜀", "-", "안 하면 捤獥汤捯 같은 깨진 문자 발생"],
    ["C. olefile-text", "3-2", "converters.hwp.olefile", "└ surrogate 쌍 결합",
     "UTF-16 상위/하위 대리를 짝지어 한 글자로", "-", "안 하면 JSON 저장 시 UnicodeEncodeError"],
    ["C. olefile-text", "3-3", "converters.hwp.olefile", "└ 사설영역 정리",
     "한글 전용 기호(PUA) 제거", "-", "다른 프로그램에서 네모로 보이는 문자"],
    ["C. olefile-text", "4", "converters.hwp.olefile", "clean_text()",
     "필드 제어 문자 정리 후 문단 구성", "없음", "표는 텍스트로만 남음"],

    ["공통", "5", "docstruct.extractors.hwp", "extract_hwp_pages()",
     "markdown → PageContent 1개 (페이지 개념 없음)", "-", "page_no_kind='whole'"],
    ["공통", "6", "docstruct.split", "split_document()",
     "구조 경계(제N장·□·번호)에서 목표 크기까지 모아 분할", "-", "split_chars 지정 시에만"],
    ["공통", "7~", "docstruct.tables.assess / fill", "(PDF 와 동일)",
     "표 판정·재추출. 근거는 원본 <table> HTML", "-",
     "olefile-text 경로는 근거가 없어 재추출 불가"],
])
style_sheet(ws, [15, 7, 27, 27, 46, 10, 40], header_row=3)
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    v = row[0].value or ""
    if v.startswith("A."):
        row[0].fill = PatternFill("solid", fgColor="E2EFDA")
    elif v.startswith("B."):
        row[0].fill = PatternFill("solid", fgColor="DDEBF7")
    elif v.startswith("C."):
        row[0].fill = PatternFill("solid", fgColor="FCE4D6")
    if (row[3].value or "").startswith("└"):
        for c in row:
            c.fill = SUB_FILL

# ──────────────────────────────────────────────────── 5. HWPX 파이프라인
ws = wb.create_sheet("10.HWPX 단계표")
add_rows(ws, [
    "단계", "모듈", "함수", "하는 일", "표 구조", "비고",
], [
    ["1", "docstruct.extractors.registry", "_extract_hwpx()",
     "확장자 .hwpx → HWPX 추출기 호출", "-", "-"],
    ["2", "docstruct.extractors.hwpx", "extract_hwpx_pages()",
     "python-hwpx 로 파싱 (import 성공 여부를 탐색보다 우선)", "-",
     "미설치 시 실행 파이썬 경로를 포함한 ImportError"],
    ["3", "hwpx (외부)", "HwpxDocument.open()",
     "zip 컨테이너 열고 section*.xml 파싱", "보존", "OOXML 계열 — PDF·HWP 보다 빠름"],
    ["4", "hwpx (외부)", "export_rich_markdown()",
     "본문·표를 markdown 으로 변환", "보존 (cellSpan)",
     "cellSpan 이 없는 비표준 파일은 라이브러리 내부 오류 발생"],
    ["5", "docstruct.tables.markdown", "inject_table_placeholders()",
     "markdown 표를 <table N> 으로 치환하고 TableInfo 생성", "-",
     "구분선 정규식이 개행을 삼키지 않도록 수정됨"],
    ["6", "docstruct.split", "split_document()",
     "구조 경계 기준 분할", "-", "split_chars 지정 시에만"],
    ["7~", "docstruct.tables.assess / fill", "(PDF 와 동일)",
     "표 판정·재추출. 근거는 원본 <table> HTML", "-", "페이지 이미지 없음"],
])
style_sheet(ws, [7, 30, 30, 46, 16, 44])

# ──────────────────────────────────────────────────── 6. 모듈 구조
ws = wb.create_sheet("11.모듈 구조")
add_rows(ws, ["계층", "패키지 / 모듈", "역할", "주요 함수·클래스", "의존"], [
    ["L0", "core.config", "설정 확정 — 환경변수 → .env → 내장 기본값",
     "get_settings() · resolve_device() · LLMEndpoint · LocalVLM", "없음"],
    ["L0", "core.winfix", "Windows 비 UTF-8 로케일 우회", "apply()", "없음"],
    ["L1", "infrastructure.llm.client", "LLM 호출 분배 — 로컬 VLM / HTTP / 대비 엔드포인트",
     "invoke_llm() · llm_available()", "core"],
    ["L1", "infrastructure.llm.local_vlm", "이 장비의 VLM 직접 실행 (transformers)",
     "invoke() · available()", "core"],
    ["L2", "converters.pdf", "Docling 어댑터 — 옵션 구성·변환·진단 수집",
     "PdfConverter · get_document_converter()", "core, docling"],
    ["L2", "converters.hwp", "HWP 3경로 (hwpml / pyhwp / olefile)",
     "HwpConverter · hwp_to_html_str() · extract_raw_text()", "core, pyhwp, olefile"],
    ["L2", "converters.html", "HTML → markdown (HWP 경로 공용)",
     "html_to_markdown()", "beautifulsoup4"],
    ["L3", "docstruct.extractors", "포맷별 추출기 레지스트리 → PageContent[]",
     "@register_extractor · get_extractor()", "converters"],
    ["L3", "docstruct.tables", "표 처리 — 변환·판정·재추출·정규화",
     "assess_document() · process_tables() · docling_table_to_markdown()", "infrastructure"],
    ["L3", "docstruct.media", "이미지 — 페이지 렌더·그림 저장·인코딩",
     "render_page_images() · save_picture() · encode_image_file()", "pypdfium2"],
    ["L3", "docstruct.models", "데이터 모델",
     "PageDocument · PageContent · TableInfo · ImageInfo · PageTrace", "없음"],
    ["L3", "docstruct.layout", "레이아웃 인식 결과 수집 (진단)",
     "collect_layout() · LayoutItem · overlapping_pairs()", "없음"],
    ["L3", "docstruct.split", "긴 문서를 구조 경계에서 분할",
     "split_document() · boundary_lines()", "models"],
    ["L4", "docstruct.pipeline", "전체 조립 — 추출 → 렌더 → 판정 → 재추출 → 출력",
     "build_document()", "L0~L3 전부"],
    ["L4", "docstruct.report", "산출물 생성 (json · md 5종)",
     "write_json() · write_markdown() · summary_lines()", "models"],
    ["L4", "docstruct.preview", "노트북 표시 (IPython 없으면 안내)",
     "show_document() · show_tables() · *_html()", "models"],
    ["L4", "docstruct.checks", "환경 점검 · 캐시 무효화",
     "environment() · invalidate_caches()", "core, converters"],
    ["L5", "docstruct.api", "공개 API",
     "DocStruct · DocStructBatch · configure() · set_model()", "L4"],
    ["L5", "docstruct.cli", "명령행", "main()", "L4"],
])
style_sheet(ws, [7, 30, 40, 48, 22])

# ──────────────────────────────────────────────────── 7. 데이터 모델
ws = wb.create_sheet("12.데이터 모델")
add_rows(ws, ["클래스", "필드", "타입", "설명"], [
    ["PageDocument", "filename", "str", "원본 파일명"],
    ["", "source_format", "str", "pdf | hwp | hwpx"],
    ["", "page_count", "int", "페이지(조각) 수 — pages 길이"],
    ["", "pages", "list[PageContent]", "페이지 목록"],
    ["", "failed_pages", "list[int]", "파싱 실패로 결과에서 빠진 페이지 번호"],
    ["", "pipeline", "dict", "이 실행에 적용된 설정 스냅샷 (llm_url·device·llm_concurrency 등)"],
    ["", "timings", "dict[str, float]", "단계별 소요 시간(초)"],

    ["PageContent", "page_no", "int", "페이지 번호"],
    ["", "page_no_kind", "str", "exact(PDF) | whole(HWP 전체) | chunk(분할됨)"],
    ["", "content", "str", "본문 markdown — 표는 <table N>, 그림은 <image N>"],
    ["", "tables", "list[TableInfo]", "이 페이지의 표"],
    ["", "images", "list[ImageInfo]", "이 페이지의 그림"],
    ["", "page_image_path", "str | None", "렌더된 페이지 PNG 경로 (PDF)"],
    ["", "trace", "PageTrace", "처리 경로 기록"],
    ["", "layout", "list[LayoutItem]", "레이아웃 인식 결과 (PDF)"],

    ["TableInfo", "id", "str", "table_1, table_2 …"],
    ["", "table_num", "int", "placeholder 번호"],
    ["", "placeholder", "str", "<table 1>"],
    ["", "markdown", "str", "GFM markdown 표"],
    ["", "original_markdown", "str | None", "재추출됐으면 원본이 여기 보관"],
    ["", "content_type", "str", "table | text | image  (LLM 판정)"],
    ["", "quality", "str", "sufficient | wrong | insufficient  (LLM 판정)"],
    ["", "llm_title", "str | None", "LLM 이 붙인 표 제목"],
    ["", "reason", "str | None", "판정 사유"],
    ["", "bbox", "dict | None", "페이지 좌표 (PDF)"],
    ["", "needs_fill", "bool (property)", "content_type=table 이고 quality 가 wrong·insufficient"],
    ["", "was_filled", "bool (property)", "재추출이 실제로 일어났는지"],

    ["ImageInfo", "id", "str", "image_1, image_2 …"],
    ["", "placeholder", "str", "<image 1>"],
    ["", "description", "str | None", "VLM 이 생성한 설명"],
    ["", "image_path", "str | None", "저장된 파일 경로"],
    ["", "mime_type", "str | None", "image/png 등"],

    ["PageTrace", "extractor", "str", "docling | hwpml-xml | pyhwp-html | olefile-text | python-hwpx"],
    ["", "text_source", "str", "text_layer | ocr | mixed | empty | unmeasured | n/a"],
    ["", "ocr_ratio", "float", "OCR 로 만들어진 셀 비율"],
    ["", "rendered / assessed / refilled", "bool", "각 단계 수행 여부"],
    ["", "steps", "list[TraceStep]", "순차 실행 기록"],
    ["", "summary()", "메서드", "한 줄 요약 — 'docling · OCR 92% · 표3 · 렌더 · 평가'"],
    ["", "log()", "메서드", "순차 실행 로그 전문"],

    ["TraceStep", "module", "str", "수행 모듈"],
    ["", "action", "str", "무엇을 했는지"],
    ["", "detail", "str", "상세"],
    ["", "status", "str", "ok | warn | skip | fail"],
    ["", "duration_ms", "float | None", "소요 시간 (LLM 단계)"],
])
style_sheet(ws, [16, 30, 20, 72])
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value:
        for c in row:
            c.fill = SUB_FILL
            c.font = Font(name=FONT, size=10, bold=True)

# ──────────────────────────────────────────────────── 8. 설정
ws = wb.create_sheet("13.설정")
add_rows(ws, ["갈래", "설정 키", "환경변수", "기본값", "설명"], [
    ["LLM", "llm_url", "DOCLING_TABLE_API_URL", "(사이트 기본값)", "표 판정·재추출·목차용 엔드포인트"],
    ["LLM", "llm_model", "DOCLING_TABLE_API_MODEL", "(사이트 기본값)", "모델명"],
    ["LLM", "llm_key", "DOCLING_TABLE_API_KEY", "-", "인증 키"],
    ["LLM", "llm_timeout", "DOCLING_TABLE_API_TIMEOUT", "120", "응답 대기(초). 연결 대기는 5초 고정"],
    ["LLM", "llm_concurrency", "DOCLING_LLM_CONCURRENCY", "4", "동시 호출 수 — I/O 대기라 코어 수와 무관"],
    ["LLM", "llm_adapter", "DOCSTRUCT_LLM_ADAPTER", "-", "외부 HTTP 어댑터 모듈명. 미지정 시 requests 직접"],
    ["대비책", "openai_key", "OPENAI_API_KEY", "-", "연결 실패 시 전환용 키. 없으면 전환 안 함"],
    ["대비책", "fallback_model", "DOCLING_TABLE_API_FALLBACK_MODEL", "gpt-5.6-luna", "대비 엔드포인트 모델"],
    ["대비책", "fallback_enabled", "DOCLING_TABLE_API_FALLBACK", "on", "off 로 끄기"],
    ["로컬 VLM", "vlm_model", "DOCSTRUCT_VLM_MODEL", "-", "지정 시 HTTP 대신 이 모델을 직접 실행"],
    ["로컬 VLM", "vlm_device", "DOCSTRUCT_VLM_DEVICE", "(device 따름)", "실행 장치"],
    ["로컬 VLM", "vlm_dtype", "DOCSTRUCT_VLM_DTYPE", "auto", "float16 | bfloat16 | float32"],
    ["로컬 VLM", "vlm_max_tokens", "DOCSTRUCT_VLM_MAX_TOKENS", "2048", "생성 상한"],
    ["대비책", "fallback_url", "DOCLING_TABLE_API_FALLBACK_URL", "https://api.openai.com/v1/chat/completions", "대비 엔드포인트 주소"],
    ["대비책", "fallback_key", "DOCLING_TABLE_API_FALLBACK_KEY", "-", "전용 키. 없으면 OPENAI_API_KEY 사용"],
    ["대비책", "fallback_timeout", "DOCLING_TABLE_API_FALLBACK_TIMEOUT", "180", "응답 대기(초)"],
    ["그림 설명", "picture_url", "DOCLING_PICTURE_API_URL", "(사이트 기본값)", "VLM 엔드포인트. docling 이 직접 호출"],
    ["그림 설명", "picture_model", "DOCLING_PICTURE_API_MODEL", "(사이트 기본값)", "VLM 모델명"],
    ["그림 설명", "picture_key", "DOCLING_PICTURE_API_KEY", "-", "인증 키"],
    ["그림 설명", "picture_enabled", "DOCLING_PICTURE_API", "on", "off 로 그림 설명 끄기"],
    ["그림 설명", "picture_area_threshold", "DOCLING_PICTURE_AREA_THRESHOLD", "0.01", "이보다 작은 그림은 호출 생략"],
    ["PDF", "pdf_backend", "DOCLING_PDF_BACKEND", "auto", "auto | pypdfium2 | dlparse"],
    ["PDF", "ocr_backend", "DOCLING_OCR_BACKEND", "rapidocr", "rapidocr | tesseract | easyocr | auto"],
    ["PDF", "ocr_lang", "DOCLING_OCR_LANG", "korean, english", "엔진별 규약 다름 (rapidocr=korean / easyocr=ko)"],
    ["PDF", "force_full_page_ocr", "DOCLING_FORCE_FULL_PAGE_OCR", "false", "텍스트 레이어 무시하고 전면 OCR"],
    ["PDF", "generate_parsed_pages", "DOCLING_GENERATE_PARSED_PAGES", "false", "페이지별 텍스트 출처 측정 (메모리 증가)"],
    ["PDF", "code_formula_enrichment", "DOCLING_CODE_FORMULA_ENRICHMENT", "false", "수식·코드 VLM. 표 추출엔 불필요"],
    ["성능", "device", "DOCLING_DEVICE", "auto", "auto | cpu | cuda | cuda:0 | mps — auto 는 쓸 수 있는 GPU 자동 탐색"],
    ["성능", "num_threads", "DOCLING_NUM_THREADS", "0", "0 이면 기본값"],
    ["성능", "threaded_pipeline", "DOCLING_THREADED_PIPELINE", "false", "Docling 단계 병렬화 (CPU 계산형)"],
    ["성능", "rapidocr_runtime", "DOCLING_RAPIDOCR_RUNTIME", "onnxruntime", "onnxruntime | torch | openvino | paddle"],
    ["실행", "assess_tables", "-", "True", "표 품질 판정 [LLM]"],
    ["실행", "fill_tables", "-", "True", "불량 표 재추출 [LLM]"],
    ["실행", "fill_all", "-", "False", "품질 무관 전체 재추출"],
    ["실행", "render_pages", "-", "True", "페이지 PNG 렌더 (PDF)"],
    ["실행", "render_scale", "-", "2.0", "렌더 배율. 1.0 = 72dpi"],
    ["실행", "split_chars", "-", "0", "0 이면 분할 안 함. 지정 시 구조 경계에서 분할"],
    ["실행", "out_dir", "-", "None", "산출물 위치. 없으면 그림은 임시 폴더"],
    ["실행", "progress", "-", "False", "진행 막대 (tqdm 없으면 로그)"],
    ["기타", "-", "DOCSTRUCT_HWP_TIMEOUT", "300", "hwp5html 제한 시간(초). 넘기면 텍스트 경로로"],
    ["기타", "-", "DOCSTRUCT_ENV", "(작업 디렉터리)", ".env 파일 위치"],
])
style_sheet(ws, [12, 26, 36, 20, 60])

# ──────────────────────────────────────────────────── 9. 산출물
ws = wb.create_sheet("14.산출물")
add_rows(ws, ["파일 / 폴더", "생성 모듈", "내용", "형식별 차이"], [
    ["document.json", "report.write_json()",
     "전체 구조 — 본문·표·이미지·처리경로·레이아웃·설정·소요시간", "공통"],
    ["document.md", "report.write_markdown()",
     "본문. placeholder 를 실제 표·그림 설명으로 펼침", "공통"],
    ["tables.md", "report.write_tables_report()",
     "표별 판정 결과 + 재추출 전/후 비교", "공통"],
    ["pipeline.md", "report.write_pipeline_report()",
     "적용 설정 · 단계별 소요 시간(GPU 영향 표시) · 페이지별 처리 경로 · 로그 전문", "공통"],
    ["layout.md", "report.write_layout_report()",
     "레이아웃 모델 라벨 vs 파이프라인 처리 결과 대조", "PDF 만 의미 있음"],
    ["outline.md", "outline.builder",
     "의미 경로(장·절)", "--outline 지정 시에만"],
    ["pages/", "media.page_render",
     "페이지 PNG (표가 있는 페이지만)", "PDF 만"],
    ["images/", "media.picture / hwp.converter",
     "추출된 그림", "PDF · pyhwp-html 경로만"],
])
style_sheet(ws, [16, 34, 66, 26])

# ──────────────────────────────────────────────────── 10. LLM 호출
ws = wb.create_sheet("15.LLM 호출")
ws["A1"] = "LLM 이 필요한 단계는 모두 선택이다. 끄거나 연결이 안 되면 파싱 결과가 그대로 나온다."
ws["A1"].font = Font(name=FONT, size=10, bold=True)
ws["A1"].fill = NOTE_FILL
ws.append([])
add_rows(ws, ["단계", "호출 주체", "입력", "출력", "호출 횟수", "실패 시"], [
    ["표 판정", "docstruct.tables.assess", "페이지 이미지 + 표 markdown",
     "content_type · quality · 제목 · 사유", "표가 있는 페이지마다 1회",
     "전부 sufficient 로 간주하고 진행"],
    ["표 재추출", "docstruct.tables.fill", "페이지 이미지(PDF) 또는 원본 <table> HTML(HWP)",
     "새 markdown 표", "needs_fill 인 표마다 1회 (fill_all 이면 전체)",
     "원본 markdown 유지"],
    ["그림 설명", "docling 내부", "그림 이미지", "캡션 문장",
     "임계값을 넘는 그림마다 1회", "사전 연결 확인 후 안 되면 캡션 생략"],
    ["목차 추출", "docstruct.outline.builder", "페이지 본문", "의미 경로",
     "페이지마다 1회", "outline.md 미생성"],
])
style_sheet(ws, [12, 28, 40, 30, 34, 34], header_row=3)

ws.append([])
ws.append(["호출 경로 (client.invoke_llm 이 분배)"])
ws.append(["1", "로컬 VLM 이 설정됨", "→ infrastructure.llm.local_vlm.invoke() — transformers 로 직접 실행"])
ws.append(["2", "DOCSTRUCT_LLM_ADAPTER 지정", "→ 그 모듈의 create_llm_adapter()"])
ws.append(["3", "기본", "→ requests 로 직접 POST (재시도 4회, Retry-After 존중)"])
ws.append(["4", "연결 불가 시", "→ 대비 엔드포인트로 전환 (키가 있을 때만). 첫 실패 후 재시도 안 함"])
for r in range(ws.max_row - 4, ws.max_row + 1):
    for c in ws[r]:
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.cell(row=ws.max_row - 4, column=1).font = Font(name=FONT, size=10, bold=True)
ws.cell(row=ws.max_row - 4, column=1).fill = SUB_FILL

# ──────────────────────────────────────────────────── 11. 진단
ws = wb.create_sheet("16.진단")
add_rows(ws, ["수단", "확인 방법", "무엇을 알 수 있나"], [
    ["처리 경로", "page.trace.summary() / page.trace.log()",
     "어느 모듈이 무엇을 했는지, 어디서 경고·생략·실패했는지"],
    ["단계별 시간", "doc.timings / ds.summary() / pipeline.md",
     "어느 구간이 느린지. '추출'이 크면 GPU, 'LLM'이 크면 llm_concurrency"],
    ["레이아웃 대조", "preview.show_layout() / layout.md",
     "표가 깨졌을 때 모델 오인식인지 변환 문제인지 구분"],
    ["파싱 실패 페이지", "doc.failed_pages",
     "Docling 이 로그로만 남기고 결과에서 조용히 빼는 페이지"],
    ["설정 스냅샷", "doc.pipeline / document.json",
     "어떤 설정으로 돌렸는지 (사후 재현용)"],
    ["환경 점검", "docstruct --check",
     "의존성 설치 상태, OCR 실행 백엔드, 연산 장치, LLM 연결"],
    ["설치 위치", "docstruct --where",
     "실행 중인 파이썬과 docstruct 버전 (업그레이드 반영 확인)"],
    ["배치 실패 목록", "batch.failures",
     "여러 문서 처리 시 어느 파일이 왜 실패했는지"],
])
style_sheet(ws, [16, 46, 72])

# ──────────────────────────────────────────────────── 저장
out = "/mnt/user-data/outputs/docstruct_정의서.xlsx"
wb.save(out)
print(f"생성: {out}")
print(f"시트 {len(wb.sheetnames)}개: {', '.join(wb.sheetnames)}")
